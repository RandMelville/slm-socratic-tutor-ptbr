"""
Monta o pacote da 2a codificacao cega (codebook v0.2), pedida pelos revisores A e C do JBCS.

O que entra: as 39 devolutivas do modelo conformante (qwen2.5:3b-instruct, 13 cenarios
x 3 reps). E exatamente o escopo pedido pelo Revisor C ("a second blind coding pass of
the model's 39 outputs against the eight mediation functions (Table 9)").

Nota sobre prevalencia: na passada do anotador 1, FM01 e 39/39 e FM05/FM06/FM08 sao
0/39. Kappa sobre variavel constante e degenerado (indefinido, ou ~0 mesmo com
concordancia total: o "paradoxo do kappa", Feinstein & Cicchetti 1990). ISSO NAO SE
RESOLVE AUMENTANDO O PACOTE: o kappa que o revisor pede e o da Tabela 9, logo tem de ser
calculado sobre estas 39. Resolve-se no relato, com concordancia bruta e um coeficiente
robusto a prevalencia (PABAK, Byrt et al. 1993; AC1, Gwet 2008) ao lado do kappa.

N_ANCORAS permite misturar devolutivas humanas ao pacote. Fica em 0 por padrao: ancoras
mediriam a confiabilidade do CODEBOOK, nao a da Tabela 9, e custariam tempo do anotador
sem responder ao que foi pedido.

Embaralhamento deterministico (seed fixa) para o pacote ser reproduzivel.
A chave (ID -> origem) fica em chave_cega.csv, que NAO vai junto com o pacote.

Saidas:
  data/segunda_codificacao_cega/codificacao_cega_v02.csv
  data/segunda_codificacao_cega/codificacao_cega_v02.xlsx
  data/segunda_codificacao_cega/chave_cega.csv          <- nao enviar ao anotador

Uso: python data/segunda_codificacao_cega/build_pacote_cego.py
"""
from pathlib import Path
import csv
import json
import random

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "data" / "segunda_codificacao_cega"

SEED = 20260727
# Cenarios ja usados na 1a rodada cega (amostra estratificada de 20 devolutivas).
CENARIOS_1A_RODADA = {1, 5, 8, 11}
N_ANCORAS = 0

FM_COLS = [f"FM0{i}" for i in range(1, 9)]


def carregar_cenarios():
    """cenario_id -> texto do aluno (slot user do prompt canonico)."""
    caminho = ROOT / "data" / "scenarios_canonical_koch.jsonl"
    textos = {}
    with caminho.open(encoding="utf-8") as fh:
        for i, linha in enumerate(fh, start=1):
            if not linha.strip():
                continue
            msgs = json.loads(linha)["messages"]
            user = next(m["content"] for m in msgs if m["role"] == "user")
            textos[i] = user.strip()
    return textos


def render_modelo(resposta_ia):
    """Devolutiva do modelo em prosa, no mesmo layout das humanas."""
    obj = json.loads(resposta_ia)
    pontos = obj.get("pontos_fortes", "")
    if isinstance(pontos, list):
        pontos = " ".join(str(p) for p in pontos)
    perguntas = obj.get("perguntas_reflexivas", []) or []
    if isinstance(perguntas, str):
        perguntas = [perguntas]
    linhas = [str(pontos).strip()]
    for n, p in enumerate(perguntas, start=1):
        linhas.append(f"{n}. {str(p).strip()}")
    return "\n\n".join(x for x in linhas if x)


def carregar_modelo():
    dados = json.loads((ROOT / "data" / "results" / "round_1_main_models.json").read_text(encoding="utf-8"))
    itens = []
    for r in dados:
        if r["modelo_testado"] != "qwen2.5:3b-instruct":
            continue
        itens.append({
            "origem": "modelo",
            "cenario": int(r["cenario_id"]),
            "rep": int(r["rep"]),
            "respondente": "qwen2.5:3b-instruct",
            "devolutiva": render_modelo(r["resposta_ia"]),
        })
    itens.sort(key=lambda x: (x["cenario"], x["rep"]))
    assert len(itens) == 39, f"esperado 39 saidas do modelo, obtido {len(itens)}"
    return itens


def carregar_humanos():
    caminho = ROOT / "data" / "baseline_humano" / "respostas_professores.jsonl"
    itens = []
    with caminho.open(encoding="utf-8") as fh:
        for linha in fh:
            if not linha.strip():
                continue
            r = json.loads(linha)
            itens.append({
                "origem": "humano",
                "cenario": int(r["cenario"]),
                "rep": 0,
                "respondente": r["respondente"],
                "devolutiva": r["devolutiva"].strip(),
            })
    assert len(itens) == 65, f"esperado 65 devolutivas humanas, obtido {len(itens)}"
    return itens


def coding_humano_v02():
    """(respondente, cenario) -> dict FM -> 0/1, da passada do anotador 1."""
    caminho = ROOT / "analises" / "codificacao_fm.csv"
    tabela = {}
    with caminho.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            chave = (r["respondente"], int(r["cenario"]))
            tabela[chave] = {c: int(r[c]) for c in FM_COLS}
    return tabela


def escolher_ancoras(humanos, coding):
    """
    Seleciona N_ANCORAS devolutivas humanas dos cenarios fora da 1a rodada, de forma
    deterministica, buscando (i) cobrir os 5 professores, (ii) espalhar pelos cenarios
    e (iii) garantir presenca das funcoes raras no corpus do modelo (FM05/FM06/FM08),
    que sao justamente as que ficariam constantes num pacote so de modelo.
    """
    elegiveis = [h for h in humanos if h["cenario"] not in CENARIOS_1A_RODADA]
    rng = random.Random(SEED)
    rng.shuffle(elegiveis)

    raras = ["FM05", "FM06", "FM08"]
    escolhidas = []
    prof_usados, cen_usados = set(), set()

    def pontua(item):
        fm = coding.get((item["respondente"], item["cenario"]), {})
        p = sum(fm.get(c, 0) for c in raras) * 4
        # FM01 e 39/39 nas saidas do modelo (o formato forca o elogio). Sem ancoras que
        # NAO reconhecem competencia, FM01 fica ~96% e seu kappa degenera igual as raras.
        p += 4 if fm.get("FM01", 1) == 0 else 0
        p += 2 if item["respondente"] not in prof_usados else 0
        p += 2 if item["cenario"] not in cen_usados else 0
        return p

    def tomar(pool):
        pool.sort(key=lambda it: (-pontua(it), it["respondente"], it["cenario"]))
        pick = pool[0]
        elegiveis.remove(pick)
        escolhidas.append(pick)
        prof_usados.add(pick["respondente"])
        cen_usados.add(pick["cenario"])

    # 1a passada: garantir que os cinco perfis de mediacao (E1-E5) estejam representados.
    # Sem isso o criterio de funcoes raras exclui o perfil mais enxuto (E3), que e
    # justamente uma das ancoras uteis para a faixa baixa de prevalencia.
    for prof in sorted({h["respondente"] for h in elegiveis}):
        if len(escolhidas) >= N_ANCORAS:
            break
        tomar([h for h in elegiveis if h["respondente"] == prof])

    # 2a passada: completa por pontuacao, ainda favorecendo cenarios nao cobertos.
    while len(escolhidas) < N_ANCORAS and elegiveis:
        tomar(list(elegiveis))

    return escolhidas


def main():
    textos = carregar_cenarios()
    modelo = carregar_modelo()
    humanos = carregar_humanos()
    coding = coding_humano_v02()

    ancoras = escolher_ancoras(humanos, coding)
    itens = modelo + ancoras

    rng = random.Random(SEED)
    rng.shuffle(itens)

    linhas_pacote, linhas_chave = [], []
    for n, it in enumerate(itens, start=1):
        rid = f"R{n:02d}"
        linhas_pacote.append({
            "ID": rid,
            "Texto do aluno": textos[it["cenario"]],
            "Devolutiva": it["devolutiva"],
            **{c: "" for c in FM_COLS},
            "MTL": "",
            "Observações": "",
        })
        linhas_chave.append({
            "ID": rid,
            "origem": it["origem"],
            "cenario": it["cenario"],
            "rep": it["rep"],
            "respondente": it["respondente"],
        })

    campos = ["ID", "Texto do aluno", "Devolutiva"] + FM_COLS + ["MTL", "Observações"]
    with (OUT / "codificacao_cega_v02.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=campos)
        w.writeheader()
        w.writerows(linhas_pacote)

    with (OUT / "chave_cega.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["ID", "origem", "cenario", "rep", "respondente"])
        w.writeheader()
        w.writerows(linhas_chave)

    escrever_xlsx(linhas_pacote, campos)

    n_mod = sum(1 for x in linhas_chave if x["origem"] == "modelo")
    n_hum = len(linhas_chave) - n_mod
    print(f"pacote: {len(linhas_pacote)} devolutivas ({n_mod} modelo + {n_hum} ancoras humanas)")
    print("professores nas ancoras:", sorted({x['respondente'] for x in linhas_chave if x['origem'] == 'humano'}))
    print("cenarios nas ancoras:", sorted({x['cenario'] for x in linhas_chave if x['origem'] == 'humano'}))
    print("decisoes a preencher:", len(linhas_pacote) * 9)


def escrever_xlsx(linhas, campos):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Codificação"

    cabecalho = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="4A3F7A")
    fundo_mtl = PatternFill("solid", fgColor="1F7A8C")

    ws.append(campos)
    for i, c in enumerate(campos, start=1):
        cel = ws.cell(row=1, column=i)
        cel.font = cabecalho
        cel.fill = fundo_mtl if c == "MTL" else fundo
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ln in linhas:
        ws.append([ln[c] for c in campos])

    # A=ID, B=texto do aluno, C=devolutiva, D..K=FM01..FM08, L=MTL, M=Observacoes
    larguras = {"A": 7, "B": 58, "C": 58, "L": 8, "M": 30}
    for col in "DEFGHIJK":
        larguras.setdefault(col, 7)
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cel in row:
            cel.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row[0].row].height = 118

    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv.prompt = "Marque 1 (presente) ou 0 (ausente). Na dúvida, 0."
    dv.promptTitle = "Presença"
    ws.add_data_validation(dv)
    dv.add(f"D2:L{ws.max_row}")  # FM01..FM08 + MTL; Observacoes (M) fica livre

    ws.freeze_panes = "D2"
    wb.save(OUT / "codificacao_cega_v02.xlsx")


if __name__ == "__main__":
    main()
